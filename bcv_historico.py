import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from pathlib import Path

def extract_date(date_string):
    """Extracts just the date from strings like 'Fecha Operacion: 27/03/2024'"""
    if isinstance(date_string, str):
        match = re.search(r'\d{2}/\d{2}/\d{4}', date_string)
        if match:
            return match.group()
    return date_string

def process_sheet(df, ws, current_row, sheet_name):
    """Process individual sheet and return updated row counter"""
    # Extract dates from B5 and D5
    b5_date = extract_date(df.iloc[4, 1] if pd.notna(df.iloc[4, 1]) else "N/A")
    d5_date = extract_date(df.iloc[4, 3] if pd.notna(df.iloc[4, 3]) else "N/A")
    
    # Process only row 15 (0-indexed row 14)
    try:
        row_data = [
            None,  # Column A (empty)
            df.iloc[14, 1],  # Column B (row 14)
            df.iloc[14, 2],  # Column C
            df.iloc[14, 3],  # Column D
            df.iloc[14, 4],  # Column E
            df.iloc[14, 5],  # Column F
            df.iloc[14, 6],  # Column G
            b5_date,  # Column H
            d5_date   # Column I
        ]
        
        # Write to worksheet
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col_num, value=value)
        
        return current_row + 1
        
    except IndexError:
        print(f"  ⚠ Row 14 not found in sheet '{sheet_name}'")
        return current_row

def merge_excel_files(input_files, output_path):
    """
    Merges multiple Excel files into one consolidated file with:
    - Standard header format
    - Values from column B (empty header)
    - Clean date formatting
    - Subtle dotted borders between files
    """
    try:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado"
        
        # ===== STYLE SETUP =====
        dotted_border = Border(bottom=Side(style='dotted'))
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        
        # ===== HEADER SETUP =====
        # Main header (merged cells)
        ws.merge_cells('D1:E1')  # "(a) Cotización M.E./US$"
        ws.merge_cells('F1:G1')  # "Bs./M.E."
        ws['D1'] = "(a) Cotización M.E./US$"
        ws['F1'] = "Bs./M.E."
        
        # Column headers
        headers = ["", "", "Moneda/País", "Compra (BID)", "Venta (ASK)", 
                  "Compra (BID)", "Venta (ASK)", "Fecha Operacion", "Fecha Valor"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            if header:
                cell.font = header_font
                cell.alignment = center_alignment
        
        # Set column widths
        for col, width in zip('ABCDEFGHI', [5, 15, 15, 15, 15, 15, 15, 15, 15]):
            ws.column_dimensions[col].width = width
        
        # ===== PROCESS ALL FILES =====
        current_row = 3
        
        for file_idx, input_file in enumerate(input_files):
            print(f"Processing file: {Path(input_file).name}")
            
            try:
                xls = pd.ExcelFile(input_file, engine='xlrd')
                
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                    current_row = process_sheet(df, ws, current_row, sheet_name)
                
                # Add dotted border after each file (except last)
                if file_idx < len(input_files) - 1:
                    for col in range(1, 10):
                        ws.cell(row=current_row-1, column=col).border = dotted_border
            
            except Exception as e:
                print(f"  ✖ Error processing {Path(input_file).name}: {str(e)}")
                continue
        
        # Save the file
        wb.save(output_path)
        print(f"\n✔ Successfully created consolidated file: {output_path}")
        print(f"• Files processed: {len(input_files)}")
        print(f"• Total rows added: {current_row-3}")
        
    except Exception as e:
        print(f"✖ Fatal error: {e}")

# Example usage:
if __name__ == "__main__":
    # Get all .xls files in current directory
    input_files = list(Path('./files').glob('*.xls'))
    
    if not input_files:
        print("No .xls files found in current directory")
    else:
        merge_excel_files(
            input_files=input_files,
            output_path="CONSOLIDATED_OUTPUT.xlsx"
        )